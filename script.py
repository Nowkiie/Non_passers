import openpyxl

wb = openpyxl.reader.excel.load_workbook(filename="list_all_abitur.xlsx")
wb.active =0
sheet = wb.active

wb1 = openpyxl.load_workbook("list_original.xlsx")
wb1.create_sheet("Лист1")
worksheet = wb1["Лист1"]


def init_exel(i:int, r:int):
    if (i == 409): return

    
    name_all_abitur = sheet['A'+str(i)].value
    view_attestat = sheet['E'+str(i)].value

    if (view_attestat == "Оригинал"):
        worksheet['A'+str(r)] = name_all_abitur
        wb1.save("list_original.xlsx")
        init_exel(i+1, r+1)
    else:
        init_exel(i+1, r)
    


if __name__ == '__main__':
    init_exel(1, 1)
