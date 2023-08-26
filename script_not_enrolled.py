import openpyxl

wb = openpyxl.reader.excel.load_workbook(filename="list_our.xlsx")
wb.active =0
sheet = wb.active

wb2 = openpyxl.reader.excel.load_workbook(filename="list_original.xlsx")
wb2.active =0
sheet1 = wb2.active

wb1 = openpyxl.load_workbook("list_our.xlsx")
wb1.create_sheet("Лист1")
worksheet = wb1["Лист1"]


def init_exel(i:int, r: int):
    if (i == 255): return

    #get name from common list
    name_people_favt = sheet1['A'+str(i)].value

    found = False
    
    #make checking name in the zachisl list 
    for z in range (0, 10):
        name_zachisl = sheet['A'+str(z+i)].value
        if (name_people_favt == name_zachisl):
            #print("found:"+name_people_favt)
            found = True
            break
    if (i > 1):
        for z in range(-i+1, i):
            name_zachisl = sheet['A'+str(i+z)].value
            if (name_people_favt == name_zachisl):
                found = True
                break

    if (i > 40):
        for z in range(0, 40):
            name_zachisl = sheet['A'+str(i-z)].value
            if (name_people_favt == name_zachisl):
                found = True
                break
        
    if (found == False):
        print("not found:"+name_people_favt)
        init_exel(i+1, r)
    else:
        init_exel(i+1, r+1)
    


if __name__ == '__main__':
    init_exel(1, 1)
